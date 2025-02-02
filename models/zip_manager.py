from models.pdf_manager import PDF
from models.dxf_manager import DXF
from models.dwg_manager import DWG
from models.group_manager import Group
from models.erro_manager import registrar_erro
import pandas as pd
import openpyxl
import rarfile
import zipfile
import shutil
import os

os.environ["PATH"] += os.pathsep + "/usr/bin"

# Processa o arquivo .zip de entrada e os 'transforma' em um arquivo.zip processado e uma planilha
class ZipFolderManager:
    
    def __init__(self, zip_folder):
        
        self.pdfs = []
        self.dxfs = []
        self.dwgs = []
        self.grupos = []
        self.folder = zip_folder
        
        # Dados extraídos da planilha de ordem de compra
        self.planilha = ' '
        self.ordem_compra =' '
        self.data_entrega = ' '
        
        # Pastas de extração e organização
        self.extraction_folder = "DESENHOS PDFs"     
        self.organization_folder = 'OrganizedFiles'
        self.organization_folder_zip = 'OrganizedFilesZip'
        self.all_pdfs = 'OrganizedFiles/PDFs'
        self.all_dwgs = 'OrganizedFiles/DWGs'
        self.all_dxfs = 'OrganizedFiles/DXFs'
        
        self.descompact_folder = ZipFolderManager.extract_folder(self) # Pasta raiz descompatacta por completo
        self.agroup_designs() # Agrupa os arquivos .pdf .dxf .dwg em objetos específicos e em seus respectivos desenhos
        self.get_ordem_compra() # Extrai a quantidade de cada peça, ordem de compra e data de entrega
        self.sheets = ZipFolderManager.create_sheet(self) # Cria a planilha com os dados extraídos
        self.organize_folders() # Organiza os arquivos em pastas de acordo com o material
        self.processed_zip = ZipFolderManager.zip_file_process(self) # Compacta as pastas organizadas (Arquivo final)
     
    # Recebe a pasta raiz .zip e realiza a primeira descompactação              
    def extract_folder(self): 
        
        try:
            # Remove a pasta de extração se ela já existir
            if os.path.exists(self.extraction_folder):
                shutil.rmtree(self.extraction_folder)

            # Cria a pasta de extração
            os.makedirs(self.extraction_folder)

            # Verifica o tipo de arquivo e realiza a extração
            
            if self.folder.lower().endswith('.zip'):
                try:
                    with zipfile.ZipFile(self.folder, 'r') as archive:
                        archive.extractall(self.extraction_folder)
                except zipfile.BadZipFile as e:
                    registrar_erro(f"Erro ao extrair ZIP: {e}")
                    return None  

            elif self.folder.lower().endswith('.rar'):
                try:
                    with rarfile.RarFile(self.folder, 'r') as archive:
                        archive.extractall(self.extraction_folder)
                except rarfile.Error as e:
                    registrar_erro(f"Erro ao extrair RAR: {e}")
                    return None 

            
            # Chama a função para extrair subarquivos
            descompact_folder = self.extract_subfiles(self.extraction_folder)

        except Exception as e:
            registrar_erro(f"Erro geral: {e}")
        
      
       
        return descompact_folder  # Retorna a pasta extraída 
    
    # Extrai subarquivos da pasta raiz
    def extract_subfiles(self, root_folder):
        
        processed_files = set()
        stack = [root_folder]

        while stack:
            current_folder = stack.pop()

            for root, dirs, files in os.walk(current_folder):
                for file in files:
                    file_path = os.path.join(root, file)

                    if file_path in processed_files:
                        continue

                    processed_files.add(file_path)

                    # Tratamento de erro para arquivos ZIP
                    if file.lower().endswith('.zip'):
                        sub_destino = os.path.join(root, os.path.splitext(file)[0])
                        os.makedirs(sub_destino, exist_ok=True)
                        try:
                            with zipfile.ZipFile(file_path, 'r') as sub_archive:
                                sub_archive.extractall(sub_destino)
                            os.remove(file_path)
                            stack.append(sub_destino)
                        except zipfile.BadZipFile as e:
                            with open("relatorio.txt", "a") as f:
                                f.write(f"O arquivo {file_path} não é um ZIP válido ou está corrompido.\n")
                                f.write(f"Erro ao processar o arquivo {file_path}:\n")
                                f.write(f"{str(e)}\n")

                    # Tratamento de erro para arquivos RAR
                    elif file.lower().endswith('.rar'):
                        sub_destino = os.path.join(root, os.path.splitext(file)[0])
                        os.makedirs(sub_destino, exist_ok=True)
                        try:
                            with rarfile.RarFile(file_path, 'r') as sub_archive:
                                sub_archive.extractall(sub_destino)
                            os.remove(file_path)
                            stack.append(sub_destino)
                        except rarfile.Error as e:
                            with open("relatorio.txt", "a") as f:
                                f.write(f"O arquivo {file_path} não é um RAR válido ou está corrompido.\n")
                                f.write(f"Erro ao processar o arquivo {file_path}:\n")
                                f.write(f"{str(e)}\n")

        return root_folder  # Retorna a pasta raiz que contém tudo

    # Percorre a pasta descompactada até encontrar arquivos .pdf .dxf .dwg e .xlsx os agrupando em objetos específicos.
    def agroup_designs(self):
        
        for root, dirs, files in os.walk(self.descompact_folder):
        
        # Caso for um diretório vazio, continua a busca
            if dirs:
                continue
        
        # Cria um grupo para cada pasta
            grupo = Group(root)
            arquivos_encontrados = False
            for file in files:
                file_path = os.path.join(root, file)

                # Cria objetos específicos para os arquivos na memsma pasta e os adiciona ao grupo
                if file.upper().endswith(".PDF"):
                    pdf_obj = PDF(file_path, os.path.basename(file))
                    self.pdfs.append(pdf_obj)
                    grupo.adicionar_arquivo(pdf_obj)
                    arquivos_encontrados = True
                    
                elif file.upper().endswith(".DXF"):
                    dxf_obj = DXF(file_path, os.path.basename(file))
                    self.dxfs.append(dxf_obj)
                    grupo.adicionar_arquivo(dxf_obj)
                    arquivos_encontrados = True
                    
                elif file.upper().endswith(".DWG"):
                    dwg_obj = DWG(file_path, os.path.basename(file))
                    self.dwgs.append(dwg_obj)
                    grupo.adicionar_arquivo(dwg_obj)
                    arquivos_encontrados = True
                
                elif file.upper().endswith(".XLSX"):
                    self.planilha = file_path                    
                
            # Adiciona o grupo se encontrar arquivos relevantes
            if arquivos_encontrados:
                self.grupos.append(grupo)
        
    # Usa a planilha de ordem de compra para extrair a quantidade de cada peça, ordem de compra e data de entrega
    def get_ordem_compra(self):
        try:
            # Carregar a planilha Excel
            df = pd.read_excel(self.planilha)

            # Função para verificar e tratar colunas
            def process_column(column_name, fill_value="Informação não disponível"):
                if column_name not in df.columns:
                    self.registrar_erro(f"A coluna '{column_name}' não foi encontrada na planilha.")
                    return None
                df[column_name] = df[column_name].fillna(fill_value)
                df[column_name] = df[column_name].apply(lambda x: str(x) if isinstance(x, (float, int)) else str(x))
                return df[column_name]

            # Processar as colunas necessárias
            self.ordem_compra = process_column("ORDEM DE COMPRA")
            self.data_entrega = process_column("DATA")

            # Verificar se as colunas obrigatórias existem
            colunas_necessarias = ["QTD", "DESCRIÇÃO", "Produto"]
            for coluna in colunas_necessarias:
                if coluna not in df.columns:
                    self.registrar_erro(f"A coluna '{coluna}' não foi encontrada na planilha.")
                    return
            
            # Processar as colunas necessárias
            df["QTD"] = process_column("QTD")
            df["DESCRIÇÃO"] = process_column("DESCRIÇÃO")
            df["Produto"] = process_column("Produto")

            # Nome do PDF a ser buscado
            for pdf in self.pdfs:
                code_group = pdf.code_group.upper() if hasattr(pdf, 'code_group') else None

                # Verifica se code_group existe para o PDF
                if code_group:
                    # Filtra as linhas onde o 'code_group' do PDF está contido na coluna "Produto"
                    linhas_filtradas = df[df["Produto"].str.contains(code_group, case=False, na=False)]

                    if not linhas_filtradas.empty:
                        # Pega a quantidade da coluna "QTD" correspondente ao código do produto
                        pdf.qtd = int(linhas_filtradas["QTD"].iloc[0])
                    else:
                        # Caso não encontre o code_group no "Produto"
                        pdf.qtd = 0  # Ou algum outro valor de fallback

        except FileNotFoundError:
            registrar_erro("Erro: O arquivo da planilha não foi encontrado.")
        except pd.errors.EmptyDataError:
            registrar_erro("Erro: A planilha está vazia ou não pôde ser lida.")
        except Exception as e:
            registrar_erro(f"Ocorreu um erro ao processar a planilha: {e}")
               
    # A função cria uma planilha com os dados extraídos dos arquivos .pdf e .dxf
    def create_sheet(self):
  
        # Nome da planilha é o mesmo do .zip da ordem de compra
        sheet_file = (os.path.splitext(self.folder)[0]) + '.xlsx'
       
        # Se o arquivo já existir, carrega-o, caso contrário cria um novo .xlsx com o cabeçalho padrão
        exist_file = os.path.exists(sheet_file)

        if exist_file:
            wb = openpyxl.load_workbook(sheet_file)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Dados"
            
            # Cabeçalho da planilha
            ws.append(["ORDEM DE COMPRA", "DATA", "LINK PDF", "CÓD. PEÇA", "DESCRIÇÃO", "QTD", "ESPESSURA", "MATERIAL", 
                    "COMPRIMENTO", "LARGURA", "ÁREA TOTAL", "ÁREA SUPERFICIAL", "TEMPO DE CORTE", "DOBRAS","REBITES", 
                    "QTD REBITES", "ROSCAS","LOTE MÍNIMO (30% DA CHAPA)", "LOTE MÁXIMO (100% DE CHAPA)", "INFORME O CAMINHO ATÉ A PASTA"])
            ws.append(["", "" ,"" ,"" ,"" ,"", "","", "", "","","","","","","", "", "","","Admin\Desktop"])
       
        # Adiciona as linhas dos arquivos .pdf e .dxf
        for grupo in self.grupos:
            grupo.process_design()
            for row in grupo.rows_to_add_first:
                ws.append(row)
            for row in grupo.rows_to_add:
                ws.append(row)
      
        # Salva a planilha
        wb.save(sheet_file)

        return sheet_file

    # Cria pastas de acordo o material de cada projeto e os move de uma forma que todos os arquivos do mesmo tipo fiquem agrupados de acordo o seu material
    def organize_folders(self):

        def create_and_move_file(file_path, file_name, file_src):
            try:
                # Cria a pasta se não existir
                if not os.path.exists(file_path):
                    os.makedirs(file_path)
                    
                    # Dá permissão total (rwx) para todos os grupos (777)
                    os.chmod(file_path, 0o777)
                    
                destino_path = os.path.join(file_path, file_name)
                
                # Remove o arquivo de destino se ele já existir
                if os.path.exists(destino_path):
                    os.remove(destino_path)
                
                # Move o arquivo para o destino
                shutil.move(file_src, destino_path)

                # Dá permissão total (rwx) para todos os grupos (777) no arquivo movido
                os.chmod(destino_path, 0o777)
                
            except OSError as e:
               registrar_erro(f"Erro ao criar ou mover o arquivo {file_name}: {e}")
            except Exception as e:
                registrar_erro(f"Ocorreu um erro inesperado ao mover {file_name}: {e}")
                
        def create_and_copy_file(file_path, file_name, file_src):
            try:
                # Cria a pasta se não existir
                if not os.path.exists(file_path):
                    os.makedirs(file_path)
                    
                    # Dá permissão total (rwx) para todos os grupos (777)
                    os.chmod(file_path, 0o777)
                    
                destino_path = os.path.join(file_path, file_name)
                
                # Remove o arquivo de destino se ele já existir
                if os.path.exists(destino_path):
                    os.remove(destino_path)
                
                # Move o arquivo para o destino
                shutil.copy(file_src, destino_path)

                # Dá permissão total (rwx) para todos os grupos (777) no arquivo movido
                os.chmod(destino_path, 0o777)
                
            except OSError as e:
                registrar_erro(f"Erro ao criar ou mover o arquivo {file_name}: {e}")
            except Exception as e:
                registrar_erro(f"Ocorreu um erro inesperado ao mover {file_name}: {e}")

        try:
           
            # Cria as pastas caso não existam
            for directory in [self.all_pdfs, self.all_dxfs, self.all_dwgs]:
                if os.path.exists(directory):
                    shutil.rmtree(directory)
                
                os.makedirs(directory)
                    
                # Dá permissão total (rwx) para todos os grupos (777)
                os.chmod(directory, 0o777)

            # Organiza os PDFs
            for pdf in self.pdfs:
                pdf_path = os.path.join(self.all_pdfs, pdf.material)
                create_and_copy_file(pdf_path, pdf.pdf_name, pdf.pdf_file)
                
                # Associa o material dos PDFs aos DXFs e DWGs correspondentes
                for dxf in self.dxfs:
                    if pdf.name in dxf.dxf_name:
                        dxf.material = pdf.material
                
                for dwg in self.dwgs:
                    if pdf.name in dwg.dwg_name:
                        dwg.material = pdf.material

            # Organiza os DXFs
            for dxf in self.dxfs:
                dxf_path = os.path.join(self.all_dxfs, dxf.material)
                create_and_move_file(dxf_path, dxf.dxf_name, dxf.dxf_file)  

            # Organiza os DWGs
            for dwg in self.dwgs:
                dwg_path = os.path.join(self.all_dwgs, dwg.material)
                create_and_move_file(dwg_path, dwg.dwg_name, dwg.dwg_file)

        except OSError as e:
            registrar_erro(f"Erro ao criar pastas ou mover arquivos: {e}")
        except Exception as e:
            registrar_erro(f"Ocorreu um erro inesperado: {e}")

    # Realiza a compactação das pastas criadas                   
    def zip_file_process(self):
        
        folder_zip = "DESENHOS PDFs.zip"
        all_zip = "ProcessedFiles.zip"
        relatorio = "relatorio.txt"
        pdfs_zip = "PDFs.zip"
        dxfs_zip = "DXFs.zip"
        dwgs_zip = "DWGs.zip"
    
        def zip_folder (base_file, zip_file):
      
            with zipfile.ZipFile(zip_file, 'w', zipfile.ZIP_DEFLATED) as zipf:
                        for root, dirs, files in os.walk(base_file):
                            for file in files:
                                caminho = os.path.join(root, file)
                                zipf.write(caminho, os.path.relpath(caminho, base_file))
                
            return zip_file
                                        
        zip_pdfs = zip_folder(self.all_pdfs, pdfs_zip)
        zip_dxfs = zip_folder(self.all_dxfs, dxfs_zip)
        zip_dwgs = zip_folder(self.all_dwgs, dwgs_zip)
        zip_extraction_folder =  zip_folder(self.extraction_folder, folder_zip)
        
        if os.path.exists(self.organization_folder_zip):
            shutil.rmtree(self.organization_folder_zip)
        
        os.makedirs(self.organization_folder_zip) 
        
        shutil.move(zip_pdfs, self.organization_folder_zip)
        shutil.move(zip_dxfs, self.organization_folder_zip)
        shutil.move(zip_dwgs, self.organization_folder_zip)
        shutil.move(self.sheets, self.organization_folder_zip)
        shutil.move(zip_extraction_folder, self.organization_folder_zip)
        
        if os.path.exists(relatorio):
            shutil.move(relatorio, self.organization_folder_zip)
            
        zip_file_processed = zip_folder(self.organization_folder_zip, all_zip)
        
        destino_path = os.path.join("uploads", zip_file_processed)
        
        if os.path.exists(destino_path):
            os.remove(destino_path)
        
        shutil.move(zip_file_processed, "uploads")

        return zip_file_processed
    
    # Limpa todas as pastas criadas
    def clean_all(self):
        def remove_folder(folder):
            try:
                # Se a pasta contiver arquivos, tente remover as permissões de leitura
                if os.path.exists(folder):
                    for root, dirs, files in os.walk(folder, topdown=False):
                        for name in files:
                            try:
                                file_path = os.path.join(root, name)
                                os.chmod(file_path, 0o777)  # Definir permissões para leitura, escrita e execução
                                os.remove(file_path)
                            except Exception as e:
                                self.registrar_erro(f"Erro ao remover arquivo {file_path}: {e}")

                        for name in dirs:
                            try:
                                dir_path = os.path.join(root, name)
                                os.chmod(dir_path, 0o777)  # Alterar permissões da pasta
                                os.rmdir(dir_path)
                            except Exception as e:
                                self.registrar_erro(f"Erro ao remover diretório {dir_path}: {e}")

                    shutil.rmtree(folder)  # Excluir a pasta
            except Exception as e:
                registrar_erro(f"Erro ao tentar remover {folder}: {e}")

        # Chame a função para limpar as pastas
        remove_folder(self.extraction_folder)
        remove_folder(self.organization_folder)
        remove_folder(self.organization_folder_zip)
        

        try:
            if os.path.exists(self.folder):
                os.remove(self.folder)
        except Exception as e:
            registrar_erro(f"Erro ao tentar remover o arquivo {self.folder}: {e}")
            
